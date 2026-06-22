from __future__ import annotations

from pathlib import Path

from fastapi.testclient import TestClient

from personal_agent.core.database import connect
from personal_agent.app import create_personal_app


def test_personal_document_artifact_exports_md_and_docx_from_current_revision(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("PERSONAL_AGENT_LLM_PROVIDER", "fake")
    db_path = tmp_path / "agent.db"
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    client = TestClient(create_personal_app(db_path, workspace))

    draft = client.post(
        "/api/personal/artifacts/drafts",
        json={
            "artifact_type": "requirement_analysis_report",
            "title": "需求分析报告",
            "content": "# 需求分析报告\n\n旧内容",
            "content_format": "markdown",
        },
    ).json()
    revised = client.post(
        f"/api/personal/artifacts/{draft['draft_uid']}/revise-manual",
        json={"content": "# 需求分析报告\n\n当前 revision 内容"},
    ).json()
    assert revised["current_revision"] == 2

    md = client.post(f"/api/personal/artifacts/{draft['draft_uid']}/export", json={"format": "md"})
    assert md.status_code == 200
    md_path = Path(md.json()["file_path"])
    assert md_path.exists()
    assert md_path.read_text(encoding="utf-8") == "# 需求分析报告\n\n当前 revision 内容"

    docx = client.post(f"/api/personal/artifacts/{draft['draft_uid']}/export", json={"format": "docx"})
    assert docx.status_code == 200
    docx_path = Path(docx.json()["file_path"])
    assert docx_path.exists()
    from docx import Document

    doc = Document(docx_path)
    assert "当前 revision 内容" in "\n".join(paragraph.text for paragraph in doc.paragraphs)

    download = client.get(f"/api/personal/artifacts/{draft['draft_uid']}/download?format=md")
    assert download.status_code == 200
    assert download.content.decode("utf-8").replace("\r\n", "\n") == "# 需求分析报告\n\n当前 revision 内容"

    with connect(db_path) as conn:
        assert conn.execute("SELECT COUNT(*) FROM artifacts").fetchone()[0] == 0


def test_personal_table_and_patch_artifacts_export_xlsx_and_diff(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("PERSONAL_AGENT_LLM_PROVIDER", "fake")
    db_path = tmp_path / "agent.db"
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    client = TestClient(create_personal_app(db_path, workspace))

    table = client.post(
        "/api/personal/artifacts/drafts",
        json={
            "artifact_type": "test_case_spec",
            "title": "测试用例规格",
            "content_format": "json_table",
            "content": '{"columns":["id","expected"],"rows":[{"id":"TC-1","expected":"通过"}]}',
        },
    ).json()
    xlsx = client.post(f"/api/personal/artifacts/{table['draft_uid']}/export", json={"format": "xlsx"})
    assert xlsx.status_code == 200
    from openpyxl import load_workbook

    workbook = load_workbook(Path(xlsx.json()["file_path"]))
    sheet = workbook.active
    assert sheet.cell(row=1, column=1).value == "id"
    assert sheet.cell(row=2, column=2).value == "通过"

    patch_text = "diff --git a/src/a.c b/src/a.c\n--- a/src/a.c\n+++ b/src/a.c\n@@ -1 +1 @@\n-old\n+new"
    patch = client.post(
        "/api/personal/artifacts/drafts",
        json={
            "artifact_type": "c_code_diff",
            "title": "C 代码 Patch",
            "content_format": "diff",
            "content": patch_text,
        },
    ).json()
    diff = client.post(f"/api/personal/artifacts/{patch['draft_uid']}/export", json={"format": "diff"})
    assert diff.status_code == 200
    diff_path = Path(diff.json()["file_path"])
    assert diff_path.exists()
    assert diff_path.read_text(encoding="utf-8") == patch_text

    unsupported = client.post(f"/api/personal/artifacts/{patch['draft_uid']}/export", json={"format": "docx"})
    assert unsupported.status_code == 400


def test_personal_artifact_open_exports_and_launches_default_app(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("PERSONAL_AGENT_LLM_PROVIDER", "fake")
    opened: list[Path] = []
    monkeypatch.setattr("personal_agent.artifact_export._open_with_default_app", lambda path: opened.append(path))
    db_path = tmp_path / "agent.db"
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    client = TestClient(create_personal_app(db_path, workspace))

    draft = client.post(
        "/api/personal/artifacts/drafts",
        json={
            "artifact_type": "requirement_analysis_report",
            "title": "可打开草稿",
            "content": "# 可打开草稿\n\n当前内容",
            "content_format": "markdown",
        },
    ).json()

    response = client.post(f"/api/personal/artifacts/{draft['draft_uid']}/open", json={"format": "md"})

    assert response.status_code == 200, response.json()
    payload = response.json()
    assert payload["status"] == "opened"
    file_path = Path(payload["file_path"])
    assert file_path.exists()
    assert opened == [file_path]
    assert file_path.read_text(encoding="utf-8") == "# 可打开草稿\n\n当前内容"
